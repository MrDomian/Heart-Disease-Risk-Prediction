import os
import numpy as np
import pickle
import openpyxl
import onnx
import onnxruntime
import pandas as pd
import seaborn as sns
import skl2onnx
import matplotlib.pyplot as plt

from onnxconverter_common import FloatTensorType
from sklearn.exceptions import NotFittedError
from sklearn.linear_model import LogisticRegression
from sklearn.model_selection import train_test_split, cross_val_predict, LeaveOneOut
from sklearn.metrics import accuracy_score, precision_score, f1_score, roc_auc_score, recall_score, confusion_matrix
from sklearn.preprocessing import OneHotEncoder, LabelEncoder, PolynomialFeatures, StandardScaler, MinMaxScaler, \
    KBinsDiscretizer
from sklearn.utils import resample
from imblearn.combine import SMOTETomek
from imblearn.over_sampling import SMOTE


# Uses the SMOTE (Synthetic Minority Over-sampling Technique) algorithm to balance data
# by generating synthetic samples for the minority class to equalize the class distribution.
def balance_data_smote(X, y):
    X_bal, y_bal = SMOTE().fit_resample(X, y)
    return X_bal, y_bal


# Combines SMOTE with Tomek Links to improve data balancing
#  by simultaneously removing excessively represented samples and generating synthetic samples.
def balance_data_smotetomek(X, y):
    X_bal, y_bal = SMOTETomek().fit_resample(X, y)
    return X_bal, y_bal


# Trains a single model on the training data and returns the prediction results for the test data.
def train_model(model, X_train, X_test, y_train):
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)
    return y_pred


# Performs cross-validation of the model on the test data and returns the prediction results.
def train_model_cross_val(model, X_train, X_test, y_train, y_test, cv):
    model.fit(X_train, y_train)
    y_pred = cross_val_predict(model, X_test, y_test, cv=cv)
    return y_pred


# Trains the model and performs Leave One Out (LOO) validation on the test data,
# returning the prediction results for each sample.
def train_model_loo(model, X_train, X_test, y_train, y_test):
    model.fit(X_train, y_train)
    loo = LeaveOneOut()
    y_pred = []
    for train_index, test_index in loo.split(X_test):
        X_train_loo, X_test_loo = X_test.values[train_index], X_test.values[test_index]
        y_train_loo, y_test_loo = y_test.values[train_index], y_test.values[test_index]
        model.fit(X_train_loo, y_train_loo)
        y_pred_loo = model.predict(X_test_loo)
        y_pred.extend(y_pred_loo)
    return y_pred


# Conducts bootstrapping, i.e., repeatedly trains the model on samples with replacement
# and returns the prediction results for each iteration.
def train_model_bootstrapping(model, X, y, num_samples, test_size):
    results = []
    for _ in range(num_samples):
        X_train, y_train = resample(X, y, replace=True, n_samples=len(X))
        X_test, y_test = resample(X, y, replace=True, n_samples=int(len(X) * test_size))

        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        results.append((y_pred, y_test))
    return results


# Performs bootstrapping using the average method, i.e., trains the model on multiple samples
# and calculates the average performance results.
def train_model_avg_bootstrapping(model, X, y, num_samples, test_size):
    results = []
    for _ in range(num_samples):
        X_train, y_train = resample(X, y, replace=True, n_samples=len(X))
        X_test, y_test = resample(X, y, replace=True, n_samples=int(len(X) * test_size))
        model.fit(X_train, y_train)
        y_pred = model.predict(X_test)
        results.append((y_pred, y_test))

    all_results = []
    for y_pred, y_test in results:
        results_list, conf_matrix = generate_results(y_pred, y_test)
        all_results.append(results_list)

    avg_results = np.mean(all_results, axis=0)
    avg_results_rounded = [round(result, 2) for result in avg_results]

    return avg_results_rounded


# Computes various performance metrics (accuracy, ROC AUC, G-Mean, etc.) based on predicted and actual values.
def generate_results(y_pred, y_test):
    accuracy = accuracy_score(y_test, y_pred)
    roc_auc = roc_auc_score(y_test, y_pred)
    f1 = f1_score(y_test, y_pred)
    sensitivity = recall_score(y_test, y_pred)
    precision = precision_score(y_test, y_pred)
    conf_matrix = confusion_matrix(y_test, y_pred)
    tn, fp, fn, tp = conf_matrix.ravel()
    g_mean = (tp / (tp + fn)) * (tn / (tn + fp))
    specificity = tn / (tn + fp)
    avg = (f1 + roc_auc + g_mean) / 3

    list_results = [round(result, 2) for result in
                    [accuracy, roc_auc, g_mean, f1, sensitivity, specificity, precision, avg]]
    return list_results, conf_matrix


# Displays the performance metrics in a readable format.
def display_results(list_results):
    accuracy, roc_auc, g_mean, f1, sensitivity, specificity, precision, avg = list_results

    print("Accuracy:", accuracy)
    print("ROC AUC Score:", roc_auc)
    print("G-Mean:", g_mean)
    print("F1 Score:", f1)
    print("Sensitivity:", sensitivity)
    print("Specificity:", specificity)
    print("Precision:", precision)
    print("AVG:", avg)


# Displays the confusion matrix as a heatmap plot.
def display_confusion_matrix(conf_matrix):
    plt.figure(figsize=(8, 6))
    sns.heatmap(conf_matrix, annot=True, cmap="PuBuGn", fmt="d")
    plt.xlabel("Predicted")
    plt.ylabel("True")
    plt.title("Confusion Matrix")
    plt.show()


# Saves the model performance results to an Excel file (.xlsx).
def save_results_to_xlsx(filename, results_list):
    try:
        wb = openpyxl.load_workbook(f'{filename}.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        sheet = wb.active
        headers = ['Model', 'Accuracy', 'ROC AUC', 'G-Mean', 'F1 Score', 'Sensitivity', 'Specificity', 'Precision', 'AVG']
        for col_num, header in enumerate(headers, 1):
            sheet.cell(row=1, column=col_num).value = header

    last_row = sheet.max_row + 1
    for row_num, result in enumerate(results_list, last_row):
        model_name, *results = result
        sheet.cell(row=row_num, column=1).value = model_name
        for col_num, value in enumerate(results, 2):
            sheet.cell(row=row_num, column=col_num).value = value

    wb.save(f'{filename}.xlsx')


# Performs one-hot encoding for categorical variables.
def one_hot_encoding(data, categorical_columns):
    encoder = OneHotEncoder()
    encoded_data = pd.DataFrame(encoder.fit_transform(data[categorical_columns]).toarray())
    encoded_data.columns = [f"{col}_{cat}" for col in categorical_columns for cat in
                            encoder.categories_[categorical_columns.index(col)]]
    data = pd.concat([data.reset_index(drop=True), encoded_data], axis=1)
    data.drop(categorical_columns, axis=1, inplace=True)
    return data


# Performs label encoding for categorical variables.
def label_encoding(data, categorical_columns):
    encoder = LabelEncoder()
    for col in categorical_columns:
        data[col] = encoder.fit_transform(data[col])
    return data


# Assigns numerical values to categorical variables based on mapping.
def ordinal_encoding(data, categorical_columns, ordinal_mapping):
    for col in categorical_columns:
        data[col] = data[col].map(ordinal_mapping[col])
    return data


# Generates polynomial feature names based on the degree.
def generate_polynomial_feature_names(data, degree):
    feature_names = [f"x0{'**' + str(d) if d > 1 else ''}{' * x' + str(i) if i > 0 else ''}"
                     for d in range(1, degree + 1) for i in range(data.shape[1])]
    return feature_names


# Creates polynomial features for the data.
def create_polynomial_features(data, degree):
    polynomial_features = PolynomialFeatures(degree=degree, include_bias=False)
    transformed_data = polynomial_features.fit_transform(data)
    new_feature_names = generate_polynomial_feature_names(data, degree)
    data = pd.DataFrame(transformed_data, columns=new_feature_names)
    return data


# Creates interaction features by multiplying pairs of features.
def create_interaction_features(data):
    interaction_data = pd.DataFrame(index=data.index)
    for i in range(len(data.columns)):
        for j in range(i + 1, len(data.columns)):
            interaction_data[data.columns[i] + '*' + data.columns[j]] = data.iloc[:, i] * data.iloc[:, j]
    data = pd.concat([data, interaction_data], axis=1)
    return data


# Standardizes the data by scaling it to have a mean of 0 and a variance of 1.
def standardize_features(data):
    scaler = StandardScaler()
    standardized_data = scaler.fit_transform(data)
    data = pd.DataFrame(standardized_data, columns=data.columns)
    return data


# Normalizes the data by scaling it to the range [0, 1].
def normalize_features(data):
    scaler = MinMaxScaler()
    normalized_data = scaler.fit_transform(data)
    data = pd.DataFrame(normalized_data, columns=data.columns)
    return data


# Discretizes numerical features into a specified number of bins.
def discretize_features(data, continuous_columns, n_bins):
    discretizer = KBinsDiscretizer(n_bins=n_bins, encode='ordinal')
    discretized_data = discretizer.fit_transform(data[continuous_columns])
    discretized_data = pd.DataFrame(discretized_data, columns=continuous_columns)
    data[continuous_columns] = discretized_data
    return data


# Removes highly correlated features based on a specified threshold.
def remove_highly_correlated_features(data, threshold):
    correlation_matrix = data.corr().abs()
    upper_triangle = correlation_matrix.where(np.triu(np.ones(correlation_matrix.shape), k=1).astype(bool))
    correlated_features = [column for column in upper_triangle.columns if any(upper_triangle[column] > threshold)]
    data.drop(correlated_features, axis=1, inplace=True)
    return data


# Removes features with a high percentage of missing data.
def remove_features_with_high_missing_data(data, missing_threshold):
    missing_data = data.isnull().mean()
    features_to_remove = missing_data[missing_data > missing_threshold].index
    data.drop(features_to_remove, axis=1, inplace=True)
    return data


# Extracts time-based features (e.g., year, month, day of the week, hour) based on a time column.
def extract_time_features(data, time_column):
    data[time_column] = pd.to_datetime(data[time_column])
    data['year'] = data[time_column].dt.year
    data['month'] = data[time_column].dt.month
    data['day'] = data[time_column].dt.day
    data['hour'] = data[time_column].dt.hour
    data['day_of_week'] = data[time_column].dt.dayofweek
    data['day_of_week_name'] = data[time_column].dt.day_name()
    data.drop(time_column, axis=1, inplace=True)
    return data


# Exports the model to the ONNX format.
def export_onnx_model(model, X_train, y_train, model_path):
    try:
        if X_train.shape[0] != y_train.shape[0]:
            raise ValueError("Number of samples in X_train and y_train must match.")
        model.fit(X_train, y_train)

        initial_type = [('float_input', FloatTensorType([None, X_train.shape[1]]))]
        onnx_model = skl2onnx.convert.convert_sklearn(model, initial_types=initial_type)

        # Create the folder if it doesn't exist
        os.makedirs(os.path.dirname(model_path), exist_ok=True)

        onnx.save_model(onnx_model, model_path)
        print("Model exported successfully.")
    except Exception as e:
        print("An error occurred during model export:", e)


# Tests the ONNX model on test data.
def test_onnx_model(model_path, test_data_path):
    try:
        session = onnxruntime.InferenceSession(model_path)
    except (FileNotFoundError, onnxruntime.OrtInvalidGraph):
        print("Error: Failed to load the ONNX model.")
        return

    try:
        test_data = np.load(test_data_path)
    except FileNotFoundError:
        print("Error: Failed to load the test data.")
        return

    input_name = session.get_inputs()[0].name
    output_name = session.get_outputs()[0].name
    predictions = session.run([output_name], {input_name: test_data})
    return predictions


# Exports the model to the pickle format.
def export_pkl_model(model, X_train, y_train, filename):
    try:
        if X_train.shape[0] != y_train.shape[0]:
            raise ValueError("Number of samples in X_train and y_train must match.")
        model.fit(X_train, y_train)

        with open(filename, 'wb') as file:
            pickle.dump(model, file)
        print("Model exported successfully.")
    except (ValueError, NotFittedError) as e:
        print("An error occurred during model export:", e)


# Tests the pickle model on test data.
def test_pkl_model(filename_pickle, filename_npy):
    try:
        with open(filename_pickle, 'rb') as file:
            model = pickle.load(file)
    except FileNotFoundError:
        print("Error: Failed to load the pickle model.")
        return

    try:
        test_data = np.load(filename_npy)
    except FileNotFoundError:
        print("Error: Failed to load the test data.")
        return

    predictions = model.predict(test_data)
    return predictions


# Exports the model to the SAV format (binary file).
def export_sav_model(model, X_train, y_train, filename):
    try:
        if X_train.shape[0] != y_train.shape[0]:
            raise ValueError("Number of samples in X_train and y_train must match.")
        model.fit(X_train, y_train)

        with open(filename, 'wb') as file:
            pickle.dump(model, file)
        print("Model exported successfully.")
    except (ValueError, NotFittedError) as e:
        print("An error occurred during model export:", e)


# Tests the SAV model on test data.
def test_sav_model(filename_pickle, filename_npy):
    try:
        with open(filename_pickle, 'rb') as file:
            model = pickle.load(file)
    except FileNotFoundError:
        print("Error: Failed to load the pickle model.")
        return

    try:
        test_data_path = filename_npy
        test_data = np.load(test_data_path)
    except FileNotFoundError:
        print("Error: Failed to load the test data.")
        return

    predictions = model.predict(test_data)
    return predictions


if __name__ == "__main__":
    data = pd.read_csv('Data/heart_disease_risk.csv')
    X = data.drop('decision', axis=1)
    y = data['decision']
    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.3)
    X_train, y_train = balance_data_smote(X_train, y_train)
    model = LogisticRegression(solver='liblinear')

    # Saving test data to numpy file
    df = np.array([
        [63.0, 1.0, 1.0, 145.0, 233.0, 1.0, 2.0, 150.0, 0.0, 2.3, 3.0, 0.0, 6.0],
        [67.0, 1.0, 4.0, 160.0, 286.0, 0.0, 2.0, 108.0, 1.0, 1.5, 2.0, 3.0, 3.0],
        [67.0, 1.0, 4.0, 120.0, 229.0, 0.0, 2.0, 129.0, 1.0, 2.6, 2.0, 2.0, 7.0],
        [37.0, 1.0, 3.0, 130.0, 250.0, 0.0, 0.0, 187.0, 0.0, 3.5, 3.0, 0.0, 3.0],
        [41.0, 0.0, 2.0, 130.0, 204.0, 0.0, 2.0, 172.0, 0.0, 1.4, 1.0, 0.0, 3.0],
    ])
    df = df.astype(np.float32)
    np.save('Data/test_data.npy', df)

    # Train model (model.predict)
    y_pred = train_model(model, X_train, X_test, y_train)
    results, conf_matrix = generate_results(y_pred, y_test)
    display_results(results)
    display_confusion_matrix(conf_matrix)
    save_results_to_xlsx('Data/results', [('Model 1', *results)])

    # Train model (cross validation)
    y_pred = train_model_cross_val(model, X_train, X_test, y_train, y_test, 2)
    results, conf_matrix = generate_results(y_pred, y_test)
    display_results(results)
    display_confusion_matrix(conf_matrix)
    save_results_to_xlsx('Data/results', [('Model 2', *results)])

    # Train model (Leave One Out)
    y_pred = train_model_loo(model, X_train, X_test, y_train, y_test)
    results, conf_matrix = generate_results(y_pred, y_test)
    display_results(results)
    display_confusion_matrix(conf_matrix)
    save_results_to_xlsx('Data/results', [('Model 3', *results)])

    # Train model (bootstrapping)
    bootstrapping_results = train_model_bootstrapping(model, X, y, num_samples=100, test_size=0.3)
    for y_pred, y_test in bootstrapping_results:
        results, conf_matrix = generate_results(y_pred, y_test)
        display_results(results)
        display_confusion_matrix(conf_matrix)
        save_results_to_xlsx('Data/results', [('Model 4', *results)])

    # Train model (bootstraping avg)
    avg_results = train_model_avg_bootstrapping(model, X, y, num_samples=1000, test_size=0.3)
    save_results_to_xlsx('Data/results', [('Model 5', *avg_results)])

    # Models export
    export_onnx_model(model, X_train, y_train, "Data/model.onnx")
    result = test_onnx_model("Data/model.onnx", "Data/test_data.npy")
    print(result)

    export_pkl_model(model, X_train, y_train, "Data/model.pkl")
    result = test_pkl_model("Data/model.pkl", "Data/test_data.npy")
    print(result)

    export_sav_model(model, X_train, y_train, "Data/model.sav")
    result = test_sav_model("Data/model.sav", "Data/test_data.npy")
    print(result)
